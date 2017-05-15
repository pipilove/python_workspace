#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
__title__ = 'python 操作excel'
__author__ = '皮'
__mtime__ = '5/16/2016-016'
__email__ = 'pipisorry@126.com'
# code is far away from bugs with the god animal protecting
    I love animals. They taste delicious.
              ┏┓      ┏┓
            ┏┛┻━━━┛┻┓
            ┃      ☃      ┃
            ┃  ┳┛  ┗┳  ┃
            ┃      ┻      ┃
            ┗━┓      ┏━┛
                ┃      ┗━━━┓
                ┃  神兽保佑    ┣┓
                ┃　永无BUG！   ┏┛
                ┗┓┓┏━┳┓┏┛
                  ┃┫┫  ┃┫┫
                  ┗┻┛  ┗┻┛
"""
FILENAME = r'C:\Users\pline\Desktop\5-3入库.XlS'


def xl_excel():
    '''
    微软官方xl模块 pip install pyvot
    '''
    import xl
    import re

    book = xl.Workbook(FILENAME)  # 打开文件
    pro_spec = book.get("H:H").get()  # 获得某一列数据
    # 对数据进行操作
    pro_spec = [item[0:item.rfind('*')] if item is not None and item.count('*') == 2 else item for item in pro_spec][1:]
    xl.view(pro_spec, to=book.get("I:I"))  # 将数据写入对应列


def xlrd_excel():
    import xlrd

    book = xlrd.open_workbook(FILENAME)
    # sheet_name = book.sheet_names()[0]  # 获得指定索引的sheet名字
    # sheet0=book.sheet_by_name(sheet_name)  #通过sheet名字来获取，当然如果你知道sheet名字了可以直接指定
    sheet0 = book.sheet_by_index(0)  # 通过sheet索引获得sheet对象


def openpyxl_excel(FILENAME=r'C:\Users\pline\Desktop\1.xlsx'):
    '''
    只能操作office 2010+版本文件
    '''
    import openpyxl as xl

    wb = xl.load_workbook(FILENAME)  # 读取excel文件
    ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])  # 选择某个sheet
    pro_spec = [i[0].value for i in ws['H1:H' + str(ws.max_row)]]  # 选择某一列的所有数据
    #  对选中的数据进行操作
    pro_spec = [item[0:item.rfind('*')] if item is not None and item.count('*') == 2 else item for item in pro_spec]
    for index, i in enumerate(ws['H1:H' + str(ws.max_row)]):
        i[0].value = pro_spec[index]  # 将数据修改到excel文件单元格中
    wb.save(FILENAME)  # 保存文件修改


xl_excel()
# xlrd_excel()
# openpyxl_excel()
